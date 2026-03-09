from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


def write_excel_log(results: list[dict], output_folder: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extraction Log"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    headers = [
        "#",
        "Source Folder",
        "MSG Filename",
        "Applicant Name",
        "Email",
        "Phone",
        "Saved Photo File(s)",
        "Status",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28

    ok_fill = PatternFill("solid", fgColor="E2EFDA")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    err_fill = PatternFill("solid", fgColor="FCE4D6")

    ok = warn = err = 0

    for row_num, result in enumerate(results, 2):
        photos = "\n".join(result["saved_paths"]) if result["saved_paths"] else "-"
        status = "OK" if result["saved_files"] else (
            "WARNING" if result["error"] == "No image attachments found" else "ERROR"
        )

        if status == "OK":
            ok += 1
            fill = ok_fill
        elif status == "WARNING":
            warn += 1
            fill = warn_fill
        else:
            err += 1
            fill = err_fill

        row_data = [
            row_num - 1,
            result["source_folder"],
            result["msg_file"],
            result["applicant_name"],
            result["email"],
            result["phone"],
            photos,
            status,
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 34
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 45
    ws.column_dimensions["H"].width = 12

    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Count"])
    ws2.append(["Total MSG files", len(results)])
    ws2.append(["Successfully extracted", ok])
    ws2.append(["Warnings (no images)", warn])
    ws2.append(["Errors", err])
    ws2.append(["Run timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    log_path = output_folder / "extraction_log.xlsx"
    wb.save(str(log_path))
    return log_path
